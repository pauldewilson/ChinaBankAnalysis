"""
Primary scraping module that extracts data to SQL Server
"""
import os
import pandas as pd
from openpyxl import load_workbook
from .server import ServerConnection
from .settings import (
    SCRAPER_VALID_FILETYPES,
    SCRAPER_TARGET_DIRECTORIES,
    SCRAPER_TOTAL_COLUMNS,
    SCRAPER_TOTAL_ROWS
)

class Scraper:
    """
    Scraper object which extracts from Excel to SQL Server
    """
    def __init__(self):
        self.target_excel_columns = self.generate_excel_columns(SCRAPER_TOTAL_COLUMNS)
        self.sql_connection = ServerConnection().engine
        self.scrape_to_sql()

    @staticmethod
    def generate_excel_columns(target_column_number:str) -> list[str]:
        """
        Returns a list of all column letters from A to target col number
        Ex: target_column_number="C" returns ['A','B','C']
        """
        columns = []
        def column_string(n):
            """
            Generates column letter from corresponding column number
            """
            string = ""
            while n > 0:
                n, remainder = divmod(n - 1, 26)
                string = chr(65 + remainder) + string
            return string

        for n in range(1, target_column_number+1):
            columns.append(column_string(n))
        return columns

    @staticmethod
    def return_workbooks(directory:str) -> list[str]:
        """
        Returns a list of all valid files from the target directory
        """
        return [f for f in os.listdir(directory)\
            if f.lower().split('.')[-1] in SCRAPER_VALID_FILETYPES]

    def scrape_to_sql(self):
        """
        Scrapes target excel docs from given directories according to set ranges
        Note: Changes here should be reflected in the landing table (sql\create_landing_table.sql)
        and within SQL Server.
        As of writing this (2021-10-23) the set target range was A1:KN1000 (300 columns, 1000 rows)
        """
        for folder in SCRAPER_TARGET_DIRECTORIES:
            # iterate within all valid workbooks within folder
            for workbook in self.return_workbooks(folder):
                # load workbook as data only (otherwise formulae are returned instead of values)
                wb = load_workbook(folder+"\\"+workbook, data_only=True)
                # create datalist for cell values
                data = []
                # iterate over every worksheet within wb
                for ws in wb.worksheets:
                    # generate target range to be extracted from each sheet (ex. A1:C10)
                    # first cell to scrape from ex. A1
                    cell_start = 'A1'
                    # last cell to scrape from ex. KN1000
                    cell_end = f'{self.target_excel_columns[-1]}{SCRAPER_TOTAL_ROWS}'
                    # ex. A1:KN1000
                    cell_range = f'{cell_start}:{cell_end}'
                    # select rows from Worksheet object
                    rows = ws[cell_range]
                    # iterate over every row and add cell value to data list
                    for row in rows:
                        data.append([cell.value for cell in row])
                    # create df of worksheet data
                    df_data = pd.DataFrame(data=data)
                    df_data.columns = self.target_excel_columns
                    # create df of metadata
                    df_meta = pd.DataFrame(
                        data={
                            'is_processed':0,
                            'upload_number':0,
                            'workbook_name':workbook,
                            'worksheet_name':ws.title,
                            'row_n':0,
                            'source_dir':folder,
                        },
                        index=[0]
                    )
                    # concat meta and data dfs
                    df_concat = pd.concat([df_meta, df_data])
                    # infill NaN cells of df_concat with metadata
                    df_concat['is_processed'] = 0
                    df_concat['upload_number'] = 0
                    df_concat['source_dir'] = folder
                    df_concat['workbook_name'] = workbook
                    df_concat['worksheet_name'] = ws.title
                    df_concat['row_n'] = list(range(df_concat.shape[0]))
                    df_concat['source_dir'] = folder
                    # remove 0th (blank generated) row
                    df_concat = df_concat.iloc[1::, ::]
                    # send df to sql
                    df_concat.to_sql(
                        name='Worksheets',
                        schema='landing',
                        index=False,
                        con=self.sql_connection,
                        if_exists='append'
                    )
                    # clear data of processed sheet
                    data = []
                # set upload number to max+1 for the given workbook
                sql_update_statement = """
                UPDATE landing.Worksheets
                SET upload_number = (
                    SELECT max(upload_number)+1
                    FROM landing.Worksheets
                )
                WHERE upload_number = 0
                """
                self.sql_connection.execute(sql_update_statement)
